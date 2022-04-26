import openpyxl
from WoodStatistics import *


def get_data(num):
    line = list(ws[f'B{num}':f'W{num}'][0])
    line = [x for x in line if x.value != '…']
    return line


def get_years(data):
    cells = [x.column_letter + '5' for x in data]
    years = ws[cells[0]:cells[-1]][0]
    years = [x.value for x in years]
    return years


def choose_line():
    s = 'Which information you would like to see? \n' \
        '1. Timber logging \n' \
        '2. Round-wood procurement'
    regions = ws['X'][6:33]
    regions = [x.value for x in regions]

    print(s)
    inf_type = int(input())

    print('Select a region of Ukraine from: ', end='')
    [print(x, end=', ') for x in regions]
    print()
    reg = input()
    line = 7 + regions.index(reg) + 33 * (inf_type - 1)

    return line


if __name__ == '__main__':
    wb = openpyxl.load_workbook('zag_der_reg_2000-2021_ue.xlsx')
    ws = wb['Shit1']

    dock_line = choose_line()

    inf = get_data(dock_line)
    riv = WoodStatistics(inf, get_years(inf))
    print(riv)
    riv.graf()

